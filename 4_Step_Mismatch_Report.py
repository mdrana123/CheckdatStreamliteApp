import win32com.client as win32
from openpyxl import Workbook

INPUT_FILE = r"C:\Users\admin\Checkdat_Stampel_AI\Data_2_Clean.xlsx"
OUTPUT_FILE = r"C:\Users\admin\Checkdat_Stampel_AI\Data_3_Mismatch_Report.xlsx"

RED_COLORINDEX = {3}  # Excel red

def bgr_int_to_rgb(color_int: int):
    b = color_int & 255
    g = (color_int >> 8) & 255
    r = (color_int >> 16) & 255
    return r, g, b

def looks_red_rgb(r, g, b):
    return r >= 160 and g <= 170 and b <= 170

def safe_value(cell):
    try:
        return cell.Value
    except Exception:
        return None

def safe_colorindex(cell):
    try:
        return int(cell.DisplayFormat.Interior.ColorIndex)
    except Exception:
        return None

def safe_color(cell):
    try:
        return int(cell.DisplayFormat.Interior.Color)
    except Exception:
        return None

# Start Excel clean
excel = win32.DispatchEx("Excel.Application")
excel.Visible = False
excel.DisplayAlerts = False
excel.ScreenUpdating = False
excel.EnableEvents = False

wb = excel.Workbooks.Open(INPUT_FILE, ReadOnly=True)
ws = wb.Worksheets(1)

# Force calculation (important for conditional formatting)
try:
    excel.CalculateFullRebuild()
except Exception:
    pass

used = ws.UsedRange
rows = used.Rows.Count
cols = used.Columns.Count

# Read header row (column titles)
headers = {}
for c in range(1, cols + 1):
    h = safe_value(used.Cells(1, c))
    if h not in (None, ""):
        headers[c] = str(h).strip()
    else:
        headers[c] = f"COL_{c}"

# Output workbook
out_wb = Workbook()
out_ws = out_wb.active
out_ws.title = "Mismatch Report"
out_ws["A1"] = "FILE"
out_ws["B1"] = "ERROR_DETAILS"

out_row = 2

for r in range(2, rows + 1):
    file_name = safe_value(used.Cells(r, 1))
    if file_name in (None, ""):
        continue

    error_parts = []

    for c in range(1, cols + 1):
        cell = used.Cells(r, c)

        is_red = False

        # 1️⃣ ColorIndex (fast, reliable for many CF rules)
        ci = safe_colorindex(cell)
        if ci in RED_COLORINDEX:
            is_red = True

        # 2️⃣ RGB fallback
        if not is_red:
            col_int = safe_color(cell)
            if col_int is not None:
                rr, gg, bb = bgr_int_to_rgb(col_int)
                if looks_red_rgb(rr, gg, bb):
                    is_red = True

        if is_red:
            val = safe_value(cell)
            if val not in (None, ""):
                col_name = headers.get(c, f"COL_{c}")
                error_parts.append(f"{col_name}: {str(val).strip()}")

    if error_parts:
        out_ws.cell(out_row, 1, file_name)
        out_ws.cell(out_row, 2, "; ".join(error_parts))
        out_row += 1

wb.Close(SaveChanges=False)
excel.Quit()

out_wb.save(OUTPUT_FILE)

print("✅ mismatch_Report.xlsx created with column titles + values")
